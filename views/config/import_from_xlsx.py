# ##############################################################################
#
#  OSIS stands for Open Student Information System. It's an application
#  designed to manage the core business of higher education institutions,
#  such as universities, faculties, institutes and professional schools.
#  The core business involves the administration of students, teachers,
#  courses, programs and so on.
#
#  Copyright (C) 2015-2025 Université catholique de Louvain (http://www.uclouvain.be)
#
#  This program is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  A copy of this license - GNU General Public License - is available
#  at the root of the source code of this program.  If not,
#  see http://www.gnu.org/licenses/.
#
# ##############################################################################
import enum
from collections import defaultdict
from io import BytesIO
from operator import itemgetter
from typing import Annotated, Any, Callable, Generic, Type

from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.utils.translation import gettext, gettext_lazy, ngettext_lazy
from django.views.generic import FormView
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import AfterValidator, ValidationError
from pydantic.main import ModelT
from pydantic_core import PydanticCustomError
from pydantic_core.core_schema import ValidationInfo

from parcours_doctoral.forms.import_from_xlsx import ImportFromXLSXForm


class ChoixOuiNon(enum.Enum):
    OUI = 'oui'
    NON = 'non'


def foreign_key_is_valid(input_value: str, info: ValidationInfo):
    """Check that the input value is associated to an existing db object."""
    if input_value not in info.context['db_data'][info.field_name]:
        raise PydanticCustomError(
            'foreign_key_value_error',
            "No related data has been found for '{value}'",
            {'value': input_value},
        )
    return input_value


ForeignKey = Annotated[str, AfterValidator(foreign_key_is_valid)]


class ValidationCondition:
    """
    Check that a row respects a specific condition.

    """

    def __init__(
        self,
        validation_method: Callable[..., bool],
        no_valid_row_message: str = '',
        valid_rows_max_number: int = None,
        too_much_valid_rows_message: str = '',
        validation_key_fields: tuple[str, ...] = None,
    ):
        """
        :param validation_method: The method evaluating the condition that the row must met.
        The row values are passed in a dictionary as the first parameter.
        :param no_valid_row_message: The message to display if no row meets the condition.
        :param valid_rows_max_number: The maximum number of rows that must met the condition.
        :param too_much_valid_rows_message: The message to display if too much rows meet the condition.
        :param validation_key_fields: If specified, the condition must be met for some sets of rows. The set is
        determined according to the values of the fields of the row (~group by).
        """
        self.validation_method = validation_method
        self.no_valid_row_message = no_valid_row_message
        self.valid_rows_max_number = valid_rows_max_number
        self.too_much_valid_rows_message = too_much_valid_rows_message
        self.get_validation_key = itemgetter(*validation_key_fields) if validation_key_fields else lambda row: 0


class UniqueIdValidationCondition(ValidationCondition):
    """Check that only one row has the same identifier."""

    def __init__(self, *fields: str):
        super().__init__(
            validation_method=lambda row: True,
            validation_key_fields=fields,
            valid_rows_max_number=1,
            too_much_valid_rows_message='A row with the same identifier has already been encountered.',
        )


class WorksheetConfig(Generic[ModelT]):
    def __init__(
        self,
        validation_model_class: Type[ModelT],
        with_header: bool,
        additional_validations_conditions: list[ValidationCondition] = None,
    ):
        self.worksheet: Worksheet | None = None
        self.validation_model_class: Type[ModelT] = validation_model_class
        self.mapping_col_name_index: dict[str, int] = {
            field: index for index, field in enumerate(self.validation_model_class.model_fields.keys())
        }
        self.with_header = with_header
        self.min_row: int = 2 if with_header else 1
        self.max_col: int = len(self.mapping_col_name_index)
        self.validation_context_data: dict[
            str | tuple[str, ...], dict[str | tuple[str, ...], int]
        ] = {}  # {'f1': {'v1': 1}} or {('f1, 'f2'): {('v1', 'v2'): 1}}

        # Validation fields
        self.validated_objects: list[ModelT] = []
        self.is_valid: bool | None = None
        self.additional_validations_conditions: list[ValidationCondition] = additional_validations_conditions or []

        self.cell_error_style = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        self.cell_default_style = PatternFill()

    def load_worksheet(self, worksheet: Worksheet):
        self.worksheet = worksheet

    def get_sets_of_data_by_col(self, *keys: str | tuple[str, ...]):
        """
        For a worksheet, retrieve a dictionary mapping the column(s) name(s) to the set of data for the specified
        columns.
        - If a key is a column name, a set of related data is associated.
        - If a key is a tuple of columns names, a set of tuple of data is associated.
        """
        set_of_data_by_key: dict[str | tuple[str, ...], set] = {}
        itemgetters_by_key: dict[str | tuple[str, ...], itemgetter] = {}

        for key in keys:
            set_of_data_by_key[key] = set()

            if isinstance(key, str):
                itemgetters_by_key[key] = itemgetter(self.mapping_col_name_index[key])
            elif isinstance(key, tuple):
                itemgetters_by_key[key] = itemgetter(*[self.mapping_col_name_index[col] for col in key])

        for row in self.worksheet.iter_rows(min_row=self.min_row, max_col=self.max_col, values_only=True):
            if all(value is None for value in row):
                continue
            for key, key_itemgetter in itemgetters_by_key.items():
                set_of_data_by_key[key].add(key_itemgetter(row))

        return set_of_data_by_key

    def validate(self):
        """
        Validate the rows of a worksheet and create a list of objects containing the valid data (one by row).
        """
        self.validated_objects = []
        self.is_valid = True

        iter_rows = self.worksheet.iter_rows(max_col=self.max_col)

        if self.with_header:
            header_row = next(iter_rows)
            columns_names = {
                field_name: header_row[field_index].value
                for field_name, field_index in self.mapping_col_name_index.items()
            }
        else:
            columns_names = {}

        cells_by_condition: dict[ValidationCondition, dict[Any, tuple[list[Cell], list[Cell]]]] = {
            condition: {} for condition in self.additional_validations_conditions
        }

        for num_row, row in enumerate(iter_rows, start=self.min_row):
            if all(cell.value is None for cell in row):
                continue

            row_as_dict: dict[str, Any] = {}
            cells_by_field_name: dict[str, Cell] = {}
            errors_by_field: dict[str, list[str]] = defaultdict(list)
            errors_types_by_field: dict[str, list[str]] = defaultdict(list)

            for field_name, field_index in self.mapping_col_name_index.items():
                cell = row[field_index]
                cells_by_field_name[field_name] = cell
                row_as_dict[field_name] = cell.value

                # Reset the style of the cell
                cell.fill = self.cell_default_style
                cell.comment = None

            try:
                self.validated_objects.append(
                    self.validation_model_class.model_validate(
                        obj=row_as_dict,
                        context={'db_data': self.validation_context_data},
                    )
                )
            except ValidationError as validation_error:
                self.is_valid = False

                for error in validation_error.errors():
                    errors_by_field[error['loc'][0]].append(error['msg'])
                    errors_types_by_field[error['loc'][0]].append(error['type'])

            # Customize the cells depending on their validity
            for field_name, errors in errors_by_field.items():
                cell = cells_by_field_name[field_name]
                cell.fill = self.cell_error_style
                cell.comment = Comment(text='\n'.join(errors), author='OSIS')
                cell.comment.errors_types = errors_types_by_field[field_name]  # For tests

            # Display the errors in the last column of the row, if there are some
            cell_with_error = self.worksheet.cell(
                row=num_row,
                column=self.max_col + 1,
                value='\n'.join(
                    f'{columns_names.get(field, field)}: ' + '; '.join(msgs) for field, msgs in errors_by_field.items()
                )
                if errors_by_field
                else '',
            )
            cell_with_error.fill = self.cell_error_style if errors_by_field else self.cell_default_style
            cell_with_error.alignment = Alignment(wrapText=True)

            # Check if the additional conditions are respected for the current row
            for condition in self.additional_validations_conditions:
                cells = cells_by_condition[condition].setdefault(condition.get_validation_key(row_as_dict), ([], []))
                cells[condition.validation_method(row_as_dict)].append(cell_with_error)

        # Check that the additional conditions are respected and display messages if it's not the case
        for condition in self.additional_validations_conditions:
            for not_matched_cells, matched_cells in cells_by_condition[condition].values():
                nb_valid_rows = len(matched_cells)

                if nb_valid_rows:
                    if condition.valid_rows_max_number and nb_valid_rows > condition.valid_rows_max_number:
                        self.is_valid = False
                        for cell in matched_cells:
                            cell.value = f'{condition.too_much_valid_rows_message}\n{cell.value}'
                            cell.fill = self.cell_error_style
                else:
                    self.is_valid = False
                    for cell in not_matched_cells:
                        cell.value = f'{condition.no_valid_row_message}\n{cell.value}'
                        cell.fill = self.cell_error_style


class ImportFromXLSXView(FormView):
    template_name = 'parcours_doctoral/config/excel_import.html'
    import_title = gettext_lazy('Import')
    form_class = ImportFromXLSXForm

    def __init__(self, worksheet_configs: tuple[WorksheetConfig, ...], *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.worksheet_configs: tuple[WorksheetConfig, ...] = worksheet_configs

    def get_context_data(self, **kwargs):
        context_data = super().get_context_data(**kwargs)
        context_data['import_title'] = self.import_title
        return context_data

    def get_success_url(self):
        return self.request.get_full_path()

    def load_worksheet_validation_context_data(self):
        """Load the worksheet validation context data if some are necessary (WorksheetConfig.validation_context_data)"""
        return

    def save_worksheet_data(self):
        """Save the worksheet data contained in the validated_objects attribute."""
        raise NotImplementedError

    def form_valid(self, form):
        response = super().form_valid(form)

        file = form.cleaned_data['file'].file

        workbook = load_workbook(filename=file)

        nb_worksheets = len(self.worksheet_configs)

        if len(workbook.worksheets) != nb_worksheets:
            form.add_error(
                'file',
                ngettext_lazy(
                    'The file must have %(nb)s worksheet.',
                    'The file must have %(nb)s worksheets.',
                    nb_worksheets,
                )
                % {'nb': nb_worksheets},
            )
            return super().form_invalid(form)

        for index, worksheet in enumerate(workbook.worksheets):
            self.worksheet_configs[index].load_worksheet(worksheet=worksheet)

        self.load_worksheet_validation_context_data()

        for worksheet_config in self.worksheet_configs:
            worksheet_config.validate()

        if all(worksheet_config.is_valid for worksheet_config in self.worksheet_configs):
            with transaction.atomic():
                self.save_worksheet_data()

            messages.success(self.request, gettext('The data have been imported.'))

            return response

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response['Content-Disposition'] = f'attachment;filename={form.cleaned_data["file"].name}'

        workbook.close()
        file.close()

        return response
